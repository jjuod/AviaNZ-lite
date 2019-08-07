
# SupportClasses.py
#
# Support classes for the AviaNZ program
# Mostly subclassed from pyqtgraph

# Version 1.5 05/08/19
# Authors: Stephen Marsland, Nirosha Priyadarshani, Julius Juodakis

#    AviaNZ birdsong analysis program
#    Copyright (C) 2017--2019

#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.

#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.

#    You should have received a copy of the GNU General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.
#     from PyQt5.QtGui import QIcon, QPixmap
from PyQt5.QtWidgets import QAbstractButton, QMessageBox
from PyQt5.QtCore import QTime, QIODevice, QBuffer, QByteArray
from PyQt5.QtMultimedia import QAudio, QAudioOutput
from PyQt5.QtGui import QPainter, QIcon, QPixmap

import pyqtgraph as pg
from pyqtgraph.Qt import QtCore, QtGui
import pyqtgraph.functions as fn

from openpyxl import load_workbook, Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font

from scipy import signal
from scipy.signal import medfilt
import SignalProc
import WaveletFunctions
import Segment

from time import sleep
import time

import librosa

import math
import numpy as np
import os, json
import sys
import re
import copy

import Wavelet
import wavio

import io

from itertools import chain, repeat
import statistics

class preProcess:
    """ This class implements few pre processing methods to avoid noise
    """
    # todo: remove duplicate preprocess in 'Wavelet Segments'

    def __init__(self,audioData=None, spInfo={}, d=False, f=True, wavelet='dmey2'):
        self.audioData = audioData
        self.spInfo = spInfo
        self.d = d  # denoise
        self.f = f  # band-pass
        self.sp = SignalProc.SignalProc([], 0, 256, 128)
        self.WaveletFunctions = WaveletFunctions.WaveletFunctions(data=self.audioData, wavelet=wavelet, maxLevel=20, samplerate=self.spInfo['SampleRate'])

    def denoise_filter(self, level=5):
        # set df=True to perform both denoise and filter
        # df=False to skip denoise
        if self.spInfo == {}:
            fs = 8000
            f1 = None
            f2 = None
        else:
            f1 = self.spInfo['FreqRange'][0]
            f2 = self.spInfo['FreqRange'][1]
            fs = self.spInfo['SampleRate']

        # Get the five level wavelet decomposition
        if self.d:
            print(level)
            denoisedData = self.WaveletFunctions.waveletDenoise(thresholdType="soft", maxLevel=level)
        else:
            denoisedData=self.audioData  # this is to avoid washing out very fade calls during the denoising

        if self.f:
            filteredDenoisedData = self.sp.ButterworthBandpass(denoisedData, fs, low=f1, high=f2)
        else:
            filteredDenoisedData = denoisedData

        return filteredDenoisedData, fs

class postProcess:
    """ This class implements few post processing methods to avoid false positives

    segments:   detected segments in form of [[s1,e1], [s2,e2],...]
    species:    species to consider
    """

    def __init__(self, audioData=None, sampleRate=0, segments=[], spInfo={}):
        self.audioData = audioData
        self.sampleRate = sampleRate
        self.segments = segments
        if spInfo != {}:
            self.minLen = spInfo['TimeRange'][0]
            if spInfo['F0']:
                self.F0 = spInfo['F0Range']
        else:
            self.minLen = 0
        # self.confirmedSegments = []  # post processed detections with confidence TP
        # self.segmentstoCheck = []  # need more testing to confirm

    def short(self, minLen=0):
        """
        Delete segments < minLen secs
        """
        if minLen == 0:
            minLen = self.minLen
        newSegments = []
        for seg in self.segments:
            if seg[0] == -1:
                newSegments.append(seg)
            elif seg[1] - seg[0] > minLen:
                newSegments.append(seg)
            else:
                continue
        self.segments = newSegments

    def wind_cal(self, data, sampleRate):
        """ Calculate wind """
        wind_lower = 2.0 * 50 / sampleRate
        wind_upper = 2.0 * 500 / sampleRate
        f, p = signal.welch(data, fs=sampleRate, window='hamming', nperseg=512, detrend=False)
        p = np.log10(p)

        limite_inf = int(round(
            p.__len__() * wind_lower))  # minimum frequency of the rainfall frequency band 0.00625
                                        # (in normalized frequency); in Hz = 0.00625 * (44100 / 2) = 100 Hz
        limite_sup = int(round(
            p.__len__() * wind_upper))  # maximum frequency of the rainfall frequency band 0.03125(in normalized
                                        # frequency); in Hz = 0.03125 * (44100 / 2) = 250 Hz
        a_wind = p[limite_inf:limite_sup]  # section of interest of the power spectral density.Step 2 in Algorithm 2.1

        return np.mean(a_wind)  # mean of the PSD in the frequency band of interest.Upper part of the step 3 in
                                # Algorithm 2.1


    def wind(self, windT=2.5, windV=0.1):
        """
        Delete wind corrupted segments, mainly wind gust
        Automatic Identification of Rainfall in Acoustic Recordings by Carol Bedoya, Claudia Isaza, Juan M.Daza, and
        Jose D.Lopez
        :param windT: wind threshold
        :param windV: the variance
        :return: None but self.segments get updated
        """
        if len(self.segments) == 0 or len(self.segments) == 1 and self.segments[0][0] == -1:
            pass
        else:
            newSegments = copy.deepcopy(self.segments)
            for seg in self.segments:
                if seg[0] == -1:
                    continue
                else:
                    data = self.audioData[int(seg[0]*self.sampleRate):int(seg[1]*self.sampleRate)]
                    n = math.ceil((len(data) / self.sampleRate))
                    window = 1
                    start = 0
                    wind = np.zeros(n)
                    for t in range(0, n, window):
                        end = min(len(data), start + window * self.sampleRate)
                        w = np.mean(self.wind_cal(data=data[start:end], sampleRate=self.sampleRate))
                        wind[t] = w
                        start += window * self.sampleRate
                    if n > 30:      # if it is at least 30 seconds check variance (of mean values over each sec)
                        print(seg, np.max(wind), statistics.variance(wind), 'n=', n)
                        if np.max(wind) > windT and statistics.variance(wind) > windV:
                            print('long and windy')
                            newSegments.remove(seg)
                    elif np.max(wind) > windT:
                        print(seg, np.max(wind), 'n=', n)
                        print('short and windy')
                        newSegments.remove(seg)
            self.segments = newSegments
        # if you want to check out the power spectrum:
        # import matplotlib.pyplot as plt
        # plt.semilogy(f, p)
        # plt.ylim([0.5e-3, 1])
        # plt.xlabel('frequency [Hz]')
        # plt.ylabel('PSD [V**2/Hz]')
        # plt.show()

    def impulse_cal(self, fs, engp=90, fp=0.75, blocksize=10):
        """
        Find sections where impulse sounds occur e.g. clicks
        window  -   window length (no overlap)
        engp    -   energy percentile (thr), the percentile of energy to inform that a section got high energy across
                    frequency bands
        fp      -   frequency percentage (thr), the percentage of frequency bands to have high energy to mark a section
                    as having impulse noise
        blocksize - max number of consecutive blocks, 10 consecutive blocks (~1/25 sec) is a good value, to not to mask
                    very close-range calls
        :return: a binary list of length len(audioData) indicating presence of impulsive noise (0) otherwise (1)
        """

        # Calculate window length
        w1 = np.floor(fs/250)      # Window length of 1/250 sec selected experimentally
        arr = [2 ** i for i in range(5, 11)]
        pos = np.abs(arr - w1).argmin()
        window = arr[pos]

        sp = SignalProc.SignalProc(self.audioData, self.sampleRate, window, window)     # No overlap
        sg = sp.spectrogram(self.audioData, multitaper=False)

        # For each frq band get sections where energy exceeds some (90%) percentile, engp
        # and generate a binary spectrogram
        sgb = np.zeros((np.shape(sg)))
        ep = np.percentile(sg, engp, axis=0)    # note thr - 90% for energy percentile
        for y in range(np.shape(sg)[1]):
            ey = sg[:, y]
            sgb[np.where(ey > ep[y]), y] = 1

        # If lots of frq bands got 1 then predict a click
        # 1 - presence of impulse noise, 0 - otherwise here
        impulse = np.where(np.count_nonzero(sgb, axis=1) > np.shape(sgb)[1] * fp, 1, 0)     # Note thr fp

        # When an impulsive noise detected, it's better to check neighbours to make sure its not a bird call
        # very close to the microphone.
        imp_inds = np.where(impulse > 0)[0].tolist()
        imp = self.countConsecutive(imp_inds, len(impulse))

        impulse = []
        for item in imp:
            if item > blocksize or item == 0:        # Note threshold - blocksize, 10 consecutive blocks ~1/25 sec
                impulse.append(1)
            else:
                impulse.append(0)

        impulse = list(chain.from_iterable(repeat(e, window) for e in impulse))  # Make it same length as self.audioData

        if len(impulse) > len(self.audioData):      # Sanity check
            impulse = impulse[0:len(self.audioData)]
        elif len(impulse) < len(self.audioData):
            gap = len(self.audioData) - len(impulse)
            impulse = np.pad(impulse, (0, gap), 'constant')

        return impulse

    def countConsecutive(self, nums, length):
        gaps = [[s, e] for s, e in zip(nums, nums[1:]) if s + 1 < e]
        edges = iter(nums[:1] + sum(gaps, []) + nums[-1:])
        edges = list(zip(edges, edges))
        edges_reps = [item[1] - item[0] + 1 for item in edges]
        res = np.zeros((length)).tolist()
        t = 0
        for item in edges:
            for i in range(item[0], item[1]+1):
                res[i] = edges_reps[t]
            t += 1
        return res

    def mergeneighbours(self, maxGap=3):
        if len(self.segments) <= 1 or len(self.segments) == 2 and self.segments[0][0] == -1:
            pass
        else:
            newSegments = copy.deepcopy(self.segments)

            meta = None
            indx = []
            chg = False
            if newSegments[0][0] == -1:
                meta = newSegments[0]
                del (newSegments[0])

            for i in range(len(newSegments) - 1):
                if newSegments[i + 1][0] - newSegments[i][1] < maxGap:
                    indx.append(i)
                    chg = True
            indx.reverse()
            for i in indx:
                newSegments[i][1] = newSegments[i + 1][1]
                del (newSegments[i + 1])

            if chg:
                if meta:
                    newSegments.insert(0, meta)
                self.segments = newSegments

    def rainClick(self):
        """
        delete random clicks e.g. rain. Check for sign of kiwi (len)
        """
        newSegments = copy.deepcopy(self.segments)
        if newSegments.__len__() > 1:
            mfcc = librosa.feature.mfcc(self.audioData, self.sampleRate)
            # Normalise
            mfcc -= np.mean(mfcc, axis=0)
            mfcc /= np.max(np.abs(mfcc), axis=0)
            mean = np.mean(mfcc[1, :])
            std = np.std(mfcc[1, :])
            thr = mean - 2 * std  # mfcc1 thr for the recording

            for seg in self.segments:
                if seg[0] == -1:
                    continue
                else:
                    secs = seg[1] - seg[0]
                    if secs > self.minLen:  # just check duration>10 sec
                        continue
                    data = self.audioData[int(seg[0]*self.sampleRate):int(seg[1]*self.sampleRate)]
                mfcc = librosa.feature.mfcc(data, self.sampleRate)
                # Normalise
                mfcc -= np.mean(mfcc, axis=0)
                mfcc /= np.max(np.abs(mfcc), axis=0)
                mfcc1 = mfcc[1, :]  # mfcc1 of the segment
                if np.min(mfcc1) < thr:
                    newSegments.remove(seg)
        self.segments = newSegments

    def fundamentalFrq(self, fileName, speciesData):
        '''
        Check for fundamental frequency of the segments, discard the segments that do not indicate the species.
        '''
        newSegments = copy.deepcopy(self.segments)
        for seg in self.segments:
            if seg[0] == -1:
                continue
            else:
                # read the sound segment and check fundamental frq.
                secs = int(seg[1] - seg[0])
                # Got to read from the source instead of using self.audioData - ff is wrong if you use self.audioData somehow
                # data = self.audioData[int(seg[0]*speciesData['SampleRate']):int(seg[1]*speciesData['SampleRate'])]
                wavobj = wavio.read(fileName, nseconds=secs, offset=seg[0])
                data = wavobj.data
                if np.shape(np.shape(data))[0] > 1:
                    data = data[:, 0]
                sampleRate = wavobj.rate
                if data is not 'float':
                    data = data.astype('float')
                if speciesData['SampleRate'] != sampleRate:
                    data = librosa.core.audio.resample(data, sampleRate, speciesData['SampleRate'])
                # denoise before fundamental frq. extraction
                sc = preProcess(audioData=data, spInfo=speciesData, d=True, f=False)  # avoid bandpass filter
                data, sampleRate = sc.denoise_filter(level=8)
                sp = SignalProc.SignalProc([], 0, 256, 128)
                sgRaw = sp.spectrogram(data, 256, 128, mean_normalise=True, onesided=True, multitaper=False)
                segment = Segment.Segment(data, sgRaw, sp, sampleRate, 256, 128)
                pitch, y, minfreq, W = segment.yin(minfreq=100)
                ind = np.squeeze(np.where(pitch > minfreq))
                pitch = pitch[ind]
                print(seg, pitch)
                if pitch.size == 0:
                    print('Segment ', seg, ' *++ no fundamental freq detected, could be faded call or noise')
                    # newSegments.remove(seg) # for now keep it
                    continue    # continue to the next seg
                if ind.size < 2:
                    if (pitch > self.F0[0]) and (pitch < self.F0[1]):
                        # print("Match with F0 of the bird, ", pitch)
                        continue
                    else:
                        print('segment ', seg, round(pitch), ' *-- fundamental freq is out of range, could be noise')
                        newSegments.remove(seg)
                elif (np.mean(pitch) > self.F0[0]) and (np.mean(pitch) < self.F0[1]):
                        continue
                else:
                    print('segment* ', seg, round(np.mean(pitch)), pitch, np.median(pitch), ' *-- fundamental freq is out of range, could be noise')
                    newSegments.remove(seg)
                    continue
        self.segments = newSegments

    # ***no use of the rest of the functions in this class for the moment.
    def eRatioConfd(self, seg, AviaNZ_extra = False):
        '''
        This is a post processor to introduce some confidence level
        high ratio --> classes 1-3 'good' calls
        low ratio --> classes 4-5 'weak' calls
        ratio = energy in band/energy above the band
        The problem with this simple classifier is that the ratio is relatively low when the
        calls are having most of the harmonics (close range)
        Mostly works
        '''
        # TODO: Check range -- species specific of course!
        # Also recording range specific -- 16KHz will be different -- resample?
        # import WaveletSegment
        # ws = WaveletSegment.WaveletSegment()
        # detected = np.where(self.detections > 0)
        # # print "det",detected
        # if np.shape(detected)[1] > 1:
        #     detected = ws.identifySegments(np.squeeze(detected))
        # elif np.shape(detected)[1] == 1:
        #     detected = ws.identifySegments(detected)
        # else:
        #     detected=[]
        if seg: # going through segments
            sp = SignalProc.SignalProc(self.audioData[int(seg[0])*self.sampleRate:int(seg[1])*self.sampleRate], self.sampleRate, 256, 128)
            self.sg = sp.spectrogram(self.audioData[int(seg[0])*self.sampleRate:int(seg[1])*self.sampleRate])
        else: # eRatio of the whole file e.g. the extracted segments
            sp = SignalProc.SignalProc(self.audioData, self.sampleRate, 256, 128)
            self.sg = sp.spectrogram(self.audioData)

        f1 = 1500
        f2 = 4000
        F1 = f1 * np.shape(self.sg)[1] / (self.sampleRate / 2.)
        F2 = f2 * np.shape(self.sg)[1] / (self.sampleRate / 2.)

        e = np.sum(self.sg[:,int(F2):],axis=1)
        eband = np.sum(self.sg[:,int(F1):int(F2)],axis=1)
        if AviaNZ_extra:
            return eband/e, 1
        else:
            return np.mean(eband/e)

    def eRatioConfdV2(self, seg):
            '''
            This is a post processor to introduce some confidence level
            testing a variation of eratio = energy in band within segment/energy in band 10sec before or after the segment
            '''
            # TODO: Check range -- species specific of course!
            # Also recording range specific -- 16KHz will be different -- resample?
            if seg:  # going through segments
                sp = SignalProc.SignalProc(self.audioData[int(seg[0]) * self.sampleRate:int(seg[1]) * self.sampleRate],
                                           self.sampleRate, 256, 128)
                self.sg = sp.spectrogram(self.audioData[int(seg[0]) * self.sampleRate:int(seg[1]) * self.sampleRate])
                # get neighbour
                if seg[0] >= 0: #10 sec before
                    sp_nbr = SignalProc.SignalProc(self.audioData[int(seg[0]-10) * self.sampleRate:int(seg[0]) * self.sampleRate],
                                           self.sampleRate, 256, 128)
                    sg_nbr = sp_nbr.spectrogram(self.audioData[int(seg[0]-10) * self.sampleRate:int(seg[0]) * self.sampleRate])
                else: # 10 sec after
                    sp_nbr = SignalProc.SignalProc(
                        self.audioData[int(seg[1]) * self.sampleRate:int(seg[1]+10) * self.sampleRate],
                        self.sampleRate, 256, 128)
                    sg_nbr = sp_nbr.spectrogram(
                        self.audioData[int(seg[1]) * self.sampleRate:int(seg[1]+10) * self.sampleRate])

            f1 = 1500
            f2 = 7000
            F1 = f1 * np.shape(self.sg)[1] / (self.sampleRate / 2.)
            F2 = f2 * np.shape(self.sg)[1] / (self.sampleRate / 2.)

            # e = np.sum(self.sg[:, int(F2):], axis=1)
            eband = np.sum(self.sg[:, int(F1):int(F2)], axis=1)
            enbr = np.sum(sg_nbr[:, int(F1):int(F2)], axis=1)
            return (np.mean(eband) / np.mean(enbr))

    def eRatioConfd2(self, thr=2.5):
        '''
        Same as above but it checks all segments (delete after Tier1)
        This is a post processor to introduce some confidence level
        high ratio --> classes 1-3 'good' calls
        low ratio --> classes 4-5 'weak' calls
        ratio = energy in band/energy above the band
        The problem with this simple classifier is that the ratio is relatively low when the
        calls are having most of the harmonics (close range)
        Mostly works
        '''
        # TODO: Check range -- species specific of course!
        # Also recording range specific -- 16KHz will be different -- resample?
        # import WaveletSegment
        # ws = WaveletSegment.WaveletSegment()
        # detected = np.where(self.detections > 0)
        # # print "det",detected
        # if np.shape(detected)[1] > 1:
        #     detected = ws.identifySegments(np.squeeze(detected))
        # elif np.shape(detected)[1] == 1:
        #     detected = ws.identifySegments(detected)
        # else:
        #     detected=[]

        sp = SignalProc.SignalProc(self.audioData, self.sampleRate, 256, 128)
        self.sg = sp.spectrogram(self.audioData)

        # f1 = 1500
        # f2 = 4000
        # F1 = f1 * np.shape(self.sg)[1] / (self.sampleRate / 2.)
        # F2 = f2 * np.shape(self.sg)[1] / (self.sampleRate / 2.)
        #
        # e = np.sum(self.sg[:,F2:],axis=1)
        # eband = np.sum(self.sg[:,F1:F2],axis=1)
        #
        # return eband/e, 1
        f1 = 1100
        f2 = 4000
        for seg in self.segments:
            # e = np.sum(self.sg[seg[0] * self.sampleRate / 128:seg[1] * self.sampleRate / 128, :]) /128     # whole frequency range
            # nBand = 128  # number of frequency bands
            e = np.sum(self.sg[seg[0] * self.sampleRate / 128:seg[1] * self.sampleRate / 128, f2 * 128 / (self.sampleRate / 2):])  # f2:
            nBand = 128 - f2 * 128 / (self.sampleRate / 2)    # number of frequency bands
            e=e/nBand   # per band power

            eBand = np.sum(self.sg[seg[0] * self.sampleRate / 128:seg[1] * self.sampleRate / 128, f1 * 128 / (self.sampleRate / 2):f2 * 128 / (self.sampleRate / 2)]) # f1:f2
            nBand = f2 * 128 / (self.sampleRate / 2) - f1 * 128 / (self.sampleRate / 2)
            eBand = eBand / nBand
            r = eBand/e
            # print seg, r
            if r>thr:
                self.confirmedSegments.append(seg)
            else:
                self.segmentstoCheck.append(seg)

    def detectClicks(self,sg=None):
        '''
        This function finds 'click' sounds that normally pick up by any detector as false positives.
        Remove those from the output.
        '''
        # TODO: this also tends to delete true positives! Try looking back and forth to see if its longer than 1 sec
        #fs = self.sampleRate
        #data = self.audioData

        if sg is None:
            sp = SignalProc.SignalProc(self.audioData, self.sampleRate, 256, 128)
            self.sg = sp.spectrogram(self.audioData)
        else:
            self.sg = sg
        # s = Segment(data, sg, sp, fs, 50)

        energy = np.sum(self.sg, axis=1)
        energy = medfilt(energy, 15)
        e2 = np.percentile(energy, 90) * 2
        # Step 1: clicks have high energy
        clicks = np.squeeze(np.where(energy > e2))
        # Step 2: clicks are short!

        # clicks = s.identifySegments(clicks, minlength=1)
        clicks = clicks * 128 / self.sampleRate  # convert frame numbers to seconds
        #c = list(set(clicks))
        #for i in c:
        #    self.detections[i] = 0        # remove clicks
        return energy, e2

class exportSegments:
    """ This class exports .data to xlsx in batch processing, review, or manual mode.
        Three different sheets are produced in the workbook:
        time stamps, presence/absence, and per second presence/absence.
        It makes the workbook if necessary.

        Inputs
            dirName:    xlsx will be stored here
            filename:   file name to be recorded inside the xlsx
            pagelen:    page length, seconds. If =datalength, entire data is treated as a single page.
            segments:   detected segments in form of [[s1,e1], [s2,e2],...]
                        OR in format [[s1, e1, fs1, fe1, sp1], [s2, e2, fs2, fe2, sp2], ...]
            species:    Species that is currently processed.
                        Species in this list will get an xlsx even if none were detected,
                        and any [s1,e1] segments will be assigned to this.
            resolution: output resolution on excel (sheet 3) in seconds. Default is 1
            batch:      if the output is coming from batch mode, will default to appending
            minLen:     minimum length of a segment in secs
            numpages:   number of pages in this file (of size pagelen)
    """

    def __init__(self, dirName, filename, pagelen, segments, species=["Don't Know"], resolution=1, batch=False, minLen=0, numpages=1):

        self.species=species
        # convert 2-col lists to 5-col lists, if needed
        self.segments = self.correctSegFormat(segments, [])

        self.dirName=dirName
        self.filename=filename
        # extract start time of the recording (in DoC format). Default is 0
        DOCRecording = re.search('(\d{6})_(\d{6})', os.path.basename(filename))
        if DOCRecording:
            startTime = DOCRecording.group(2)
            self.startTime = int(startTime[:2])*3600 + int(startTime[2:4])*60 + int(startTime[4:6])
        else:
            self.startTime = 0
        self.numpages=numpages
        self.pagelen = math.ceil(pagelen)
        self.resolution = resolution
        self.minLen = minLen
        self.batch = batch

    def correctSegFormat(self, seglist, species):
        # Checks and if needed corrects 2-col segments to 5-col segments.
        if len(seglist)>0:
            if len(seglist[0])==2:
                print("Using old format segment list")
                # convert to new format
                for seg in seglist:
                    seg.append(0)
                    seg.append(0)
                    seg.append(species)
                return(seglist)
            elif len(seglist[0])==5:
                # using new format segment list
                return(seglist)
            else:
                print("ERROR: incorrect segment format")
                return
        else:
            return([])

    def makeNewWorkbook(self, species):
        self.wb = Workbook()
        self.wb.create_sheet(title='Time Stamps', index=1)
        self.wb.create_sheet(title='Presence Absence', index=2)
        self.wb.create_sheet(title='Per Time Period', index=3)

        ws = self.wb['Time Stamps']
        ws.cell(row=1, column=1, value="File Name")
        ws.cell(row=1, column=2, value="start (hh:mm:ss)")
        ws.cell(row=1, column=3, value="end (hh:mm:ss)")
        ws.cell(row=1, column=4, value="min freq. (Hz)")
        ws.cell(row=1, column=5, value="max freq. (Hz)")
        if species=="All_species":
            ws.cell(row=1, column=6, value="species")

        # Second sheet
        ws = self.wb['Presence Absence']
        ws.cell(row=1, column=1, value="File Name")
        ws.cell(row=1, column=2, value="Presence/Absence")

        # Third sheet
        ws = self.wb['Per Time Period']
        ws.cell(row=1, column=1, value="File Name")
        ws.cell(row=1, column=2, value="Page")
        ws.cell(row=1, column=3, value="Presence=1, Absence=0")

        # Hack to delete original sheet
        del self.wb['Sheet']
        return self.wb

    def excel(self):
        """ This saves the detections in three different formats: time stamps, presence/absence, and per second presence/absence in an excel workbook. It makes the workbook if necessary.
        Saves each species into a separate workbook,
        + an extra workbook for all species (to function as a readable segment printout).
        """
        # identify all unique species
        speciesList = set()
        for sp in self.species:
            speciesList.add(sp)
        for seg in self.segments:
            if seg[0]==-1:
                continue
            for birdName in seg[4]:
                segmentSpecies = birdName
                if birdName.endswith('?'):
                    segmentSpecies = segmentSpecies[:-1]
                speciesList.add(segmentSpecies)
        speciesList.add("All species")
        print("The following species were detected for export:")
        print(speciesList)

        def writeToExcelp1(segments):
            ws = wb['Time Stamps']
            r = ws.max_row + 1
            # Print the filename
            ws.cell(row=r, column=1, value=relfname)
            # Loop over the segments
            for seg in segments:
                ws.cell(row=r, column=2, value=str(QTime(0,0,0).addSecs(seg[0]+self.startTime).toString('hh:mm:ss')))
                ws.cell(row=r, column=3, value=str(QTime(0,0,0).addSecs(seg[1]+self.startTime).toString('hh:mm:ss')))
                if seg[3]!=0:
                    ws.cell(row=r, column=4, value=int(seg[2]))
                    ws.cell(row=r, column=5, value=int(seg[3]))
                if species=="All species":
                    ws.cell(row=r, column=6, value=", ".join(seg[4]))
                r += 1

        def writeToExcelp2(segments):
            ws = wb['Presence Absence']
            r = ws.max_row + 1
            ws.cell(row=r, column=1, value=relfname)
            ws.cell(row=r, column=2, value='_')
            if len(segments)>0:
                # if seg[1]-seg[0] > self.minLen: # skip very short segments
                ws.cell(row=r, column=2, value='Yes')
                # break
            else:
                ws.cell(row=r, column=2, value='No')

        def writeToExcelp3(detected, pagenum):
            # writes binary output DETECTED (per s) from page PAGENUM of length SELF.PAGELEN
            starttime = pagenum * self.pagelen
            ws = wb['Per Time Period']
            # print resolution "header"
            r = ws.max_row + 1
            ws.cell(row=r, column=1, value=str(self.resolution) + ' secs resolution')
            ft = Font(color=colors.DARKYELLOW)
            ws.cell(row=r, column=1).font=ft
            # print file name and page number
            ws.cell(row=r+1, column=1, value=relfname)
            ws.cell(row=r+1, column=2, value=str(pagenum+1))
            # fill the header and detection columns
            c = 3
            for t in range(0, len(detected), self.resolution):
                # absolue (within-file) times:
                win_start = starttime + t
                win_end = min(win_start+self.resolution, int(self.pagelen * self.numpages))
                ws.cell(row=r, column=c, value=str(win_start) + '-' + str(win_end))
                ws.cell(row=r, column=c).font = ft
                # within-page times:
                det = 1 if np.sum(detected[t:win_end-starttime])>0 else 0
                ws.cell(row=r+1, column=c, value=det)
                c += 1

        # now, generate the actual files, SEPARATELY FOR EACH SPECIES:
        for species in speciesList:
            print("Exporting species %s" % species)
            # clean version for filename
            speciesClean = re.sub(r'\W', "_", species)

            # setup output files:
            # if an Excel exists, append (so multiple files go into one worksheet)
            # if not, create new
            self.eFile = self.dirName + '/DetectionSummary_' + speciesClean + '.xlsx'

            if os.path.isfile(self.eFile):
                if self.batch:
                    # must not throw a dialog for every file in batch mode!
                    action = "append"
                else:
                    # check with user
                    msg = MessagePopup("w", "Excel file exists", "Output file " + self.eFile + " exists. Overwrite it, append to it, or cancel the operation?")
                    msg.setStandardButtons(QMessageBox.Cancel)
                    msg.addButton("Overwrite", QMessageBox.YesRole)
                    msg.addButton("Append", QMessageBox.YesRole)
                    # cancelBtn = msg.addButton(QMessageBox.Cancel)
                    reply = msg.exec_()
                    print(reply)
                    if reply == 4194304:  # weird const for Cancel
                        return 0
                    elif reply == 1:
                        action = "append"
                    elif reply == 0:
                        action = "overwrite"
                    else:
                        print("ERROR: Unrecognized reply", reply)
                        return 0
            else:
                # well make new book actually
                action = "overwrite"

            if action == "append":
                try:
                    wb = load_workbook(self.eFile)
                except Exception as e:
                    print("ERROR: cannot open file %s to append" % self.eFile)  # no read permissions or smth
                    print(e)
                    return 0
            elif action == "overwrite":
                try:
                    wb = self.makeNewWorkbook(speciesClean)
                except Exception as e:
                    print("ERROR: could not create new file %s" % self.eFile)  # no read permissions or smth
                    print(e)
                    return 0
            else:
                print("ERROR: unrecognized action", action)
                return 0

            try:
                relfname = str(os.path.relpath(str(self.filename), str(self.dirName)))
            except Exception as e:
                print("Falling back to absolute paths. Encountered exception:")
                print(e)
                relfname = str(os.path.abspath(str(self.filename)))
            # extract SINGLE-SPECIES ONLY segments,
            # incl. potential assignments ('Kiwi?').
            # if species=="All", take ALL segments.
            segmentsWPossible = []
            for seg in self.segments:
                if seg[0] == -1:
                    continue
                if len(seg) == 2:
                    seg.append(0)
                    seg.append(0)
                    seg.append(species)
                if species in seg[4] or species+'?' in seg[4] or species == "All species":
                    segmentsWPossible.append(seg)
            # if len(segmentsWPossible)==0:
            #     print("Warning: no segments found for species %s" % species)
            #     continue

            # export segments
            writeToExcelp1(segmentsWPossible)
            # export presence/absence
            writeToExcelp2(segmentsWPossible)

            # Generate per second binary output
            # (assuming all pages are of same length as current data)
            for p in range(0, self.numpages):
                detected = np.zeros(self.pagelen)
                for seg in segmentsWPossible:
                    for t in range(self.pagelen):
                        # convert within-page time to segment (within-file) time
                        truet = t + p*self.pagelen
                        if math.floor(seg[0]) <= truet and truet < math.ceil(seg[1]):
                            detected[t] = 1
                # write page p to xlsx
                writeToExcelp3(detected, p)

            # Save the file
            wb.save(self.eFile)
        return 1

class TimeAxisHour(pg.AxisItem):
    # Time axis (at bottom of spectrogram)
    # Writes the time as hh:mm:ss, and can add an offset
    def __init__(self, *args, **kwargs):
        super(TimeAxisHour, self).__init__(*args, **kwargs)
        self.offset = 0
        self.setLabel('Time', units='hh:mm:ss')

    def tickStrings(self, values, scale, spacing):
        # Overwrite the axis tick code
        return [QTime(0,0,0).addSecs(value+self.offset).toString('hh:mm:ss') for value in values]

    def setOffset(self,offset):
        self.offset = offset
        #self.update()

class TimeAxisMin(pg.AxisItem):
    # Time axis (at bottom of spectrogram)
    # Writes the time as mm:ss, and can add an offset
    def __init__(self, *args, **kwargs):
        super(TimeAxisMin, self).__init__(*args, **kwargs)
        self.offset = 0
        self.setLabel('Time', units='mm:ss')

    def tickStrings(self, values, scale, spacing):
        # Overwrite the axis tick code
        return [QTime(0,0,0).addSecs(value+self.offset).toString('mm:ss') for value in values]

    def setOffset(self,offset):
        self.offset = offset
        self.update()

class TimeAxisSec(pg.AxisItem):
    # Time axis (at bottom of spectrogram)
    # Writes the time as mm:ss, and can add an offset
    def __init__(self, *args, **kwargs):
        super(TimeAxisSec, self).__init__(*args, **kwargs)
        self.offset = 0
        self.setLabel('Time', units='s')

    def tickStrings(self, values, scale, spacing):
        # Overwrite the axis tick code
        return [QTime(0,0,0).addSecs(value+self.offset).toString('s') for value in values]

    def setOffset(self,offset):
        self.offset = offset
        self.update()

class FixedLineROI(pg.LineSegmentROI):
    def clearHandles(self):
        self.scene().removeItem(self.handles[0]['item'])
        self.scene().removeItem(self.handles[1]['item'])
        #while len(self.handles) > 0:
        #    self.removeHandle(self.handles[0]['item'])

class ShadedROI(pg.ROI):
    # A region of interest that is shaded, for marking segments
    def paint(self, p, opt, widget):
        #brush = QtGui.QBrush(QtGui.QColor(0, 0, 255, 50))
        if not hasattr(self, 'currentBrush'):
            self.setBrush(QtGui.QBrush(QtGui.QColor(0, 0, 255, 50)))
        if not hasattr(self, 'currentPen'):
            self.setPen(QtGui.QPen(QtGui.QColor(255, 0, 0, 255)))
        p.save()
        r = self.boundingRect()
        p.setRenderHint(QtGui.QPainter.Antialiasing)
        p.setPen(self.currentPen)
        p.setBrush(self.currentBrush)

        p.translate(r.left(), r.top())
        p.scale(r.width(), r.height())
        p.drawRect(0, 0, 1, 1)
        p.restore()

    def setBrush(self, *br, **kargs):
        """Set the brush that fills the region. Can have any arguments that are valid
        for :func:`mkBrush <pyqtgraph.mkBrush>`.
        """
        self.brush = fn.mkBrush(*br, **kargs)
        self.currentBrush = self.brush

    # this allows compatibility with LinearRegions:
    def setHoverBrush(self, *br, **kargs):
        self.hoverBrush = fn.mkBrush(*br, **kargs)

    def setPen(self, *br, **kargs):
        self.pen = fn.mkPen(*br, **kargs)
        self.currentPen = self.pen

    def hoverEvent(self, ev):
        if self.transparent:
            return
        if not ev.isExit():
            self.setMouseHover(True)
        else:
            self.setMouseHover(False)

    def setMouseHover(self, hover):
        # for ignoring when ReadOnly enabled:
        if not self.translatable:
            return
        # don't waste time if state isn't changing:
        if self.mouseHovering == hover:
            return
        self.mouseHovering = hover
        if hover:
            self.currentBrush = self.hoverBrush
        else:
            self.currentBrush = self.brush
        self.update()

def mouseDragEventFlexible(self, ev):
    if ev.button() == self.rois[0].parent.MouseDrawingButton:
        return
    ev.accept()
    
    ## Inform ROIs that a drag is happening 
    ##  note: the ROI is informed that the handle has moved using ROI.movePoint
    ##  this is for other (more nefarious) purposes.
    #for r in self.roi:
        #r[0].pointDragEvent(r[1], ev)
        
    if ev.isFinish():
        if self.isMoving:
            for r in self.rois:
                r.stateChangeFinished()
        self.isMoving = False
    elif ev.isStart():
        for r in self.rois:
            r.handleMoveStarted()
        self.isMoving = True
        self.startPos = self.scenePos()
        self.cursorOffset = self.scenePos() - ev.buttonDownScenePos()
        
    if self.isMoving:  ## note: isMoving may become False in mid-drag due to right-click.
        pos = ev.scenePos() + self.cursorOffset
        self.movePoint(pos, ev.modifiers(), finish=False)

def mouseDragEventFlexibleLine(self, ev):
    if self.movable and ev.button() != self.btn:
        if ev.isStart():
            self.moving = True
            self.cursorOffset = self.pos() - self.mapToParent(ev.buttonDownPos())
            self.startPosition = self.pos()
        ev.accept()

        if not self.moving:
            return

        self.setPos(self.cursorOffset + self.mapToParent(ev.pos()))
        self.sigDragged.emit(self)
        if ev.isFinish():
            self.moving = False
            self.sigPositionChangeFinished.emit(self)

class ShadedRectROI(ShadedROI):
    # A rectangular ROI that it shaded, for marking segments
    def __init__(self, pos, size, centered=False, movable=True, sideScalers=False, parent=None, **args):
        #QtGui.QGraphicsRectItem.__init__(self, 0, 0, size[0], size[1])
        pg.ROI.__init__(self, pos, size, movable=movable, **args)
        self.parent = parent
        self.mouseHovering = False
        self.setBrush(QtGui.QBrush(QtGui.QColor(0, 0, 255, 50)))
        self.setHoverBrush(QtGui.QBrush(QtGui.QColor(0, 0, 255, 100)))
        self.transparent = True
        if centered:
            center = [0.5, 0.5]
        else:
            center = [0, 0]

        #self.addTranslateHandle(center)
        if self.translatable:
            self.addScaleHandle([1, 1], center)
            if sideScalers:
                self.addScaleHandle([1, 0.5], [center[0], 0.5])
                self.addScaleHandle([0.5, 1], [0.5, center[1]])

    def setMovable(self,value):
        self.resizable = value
        self.translatable = value

    def mouseDragEvent(self, ev):
        if ev.isStart():
            if ev.button() != self.parent.MouseDrawingButton:
                self.setSelected(True)
                if self.translatable:
                    self.isMoving = True
                    self.preMoveState = self.getState()
                    self.cursorOffset = self.pos() - self.mapToParent(ev.buttonDownPos())
                    self.sigRegionChangeStarted.emit(self)
                    ev.accept()
                else:
                    ev.ignore()

        elif ev.isFinish():
            if self.translatable:
                if self.isMoving:
                    self.stateChangeFinished()
                self.isMoving = False
            return

        if self.translatable and self.isMoving and ev.buttons() != self.parent.MouseDrawingButton:
            snap = True if (ev.modifiers() & QtCore.Qt.ControlModifier) else None
            newPos = self.mapToParent(ev.pos()) + self.cursorOffset
            self.translate(newPos - self.pos(), snap=snap, finish=False)

pg.graphicsItems.ROI.Handle.mouseDragEvent = mouseDragEventFlexible
pg.graphicsItems.InfiniteLine.InfiniteLine.mouseDragEvent = mouseDragEventFlexibleLine

class LinearRegionItem2(pg.LinearRegionItem):
    def __init__(self, parent, *args, **kwds):
        pg.LinearRegionItem.__init__(self, *args, **kwds)
        self.parent = parent
        self.lines[0].btn = self.parent.MouseDrawingButton
        self.lines[1].btn = self.parent.MouseDrawingButton

    def mouseDragEvent(self, ev):
        if not self.movable or ev.button()==self.parent.MouseDrawingButton:
            return
        ev.accept()

        if ev.isStart():
            bdp = ev.buttonDownPos()
            self.cursorOffsets = [l.pos() - bdp for l in self.lines]
            self.startPositions = [l.pos() for l in self.lines]
            self.moving = True

        if not self.moving:
            return

        self.lines[0].blockSignals(True)  # only want to update once
        for i, l in enumerate(self.lines):
            l.setPos(self.cursorOffsets[i] + ev.pos())
        self.lines[0].blockSignals(False)
        self.prepareGeometryChange()

        if ev.isFinish():
            self.moving = False
            self.sigRegionChangeFinished.emit(self)
        else:
            self.sigRegionChanged.emit(self)


class DragViewBox(pg.ViewBox):
    # A normal ViewBox, but with the ability to capture drag.
    # Effectively, if "dragging" is enabled, it captures press & release signals.
    # Otherwise it ignores the event, which then goes to the scene(),
    # which only captures click events.
    sigMouseDragged = QtCore.Signal(object,object,object)
    keyPressed = QtCore.Signal(int)

    def __init__(self, parent, enableDrag, thisIsAmpl, *args, **kwds):
        pg.ViewBox.__init__(self, *args, **kwds)
        self.enableDrag = enableDrag
        self.parent = parent
        self.thisIsAmpl = thisIsAmpl

    def mouseDragEvent(self, ev):
        print("Uncaptured drag event")
        # if self.enableDrag:
        #     ## if axis is specified, event will only affect that axis.
        #     ev.accept()
        #     if self.state['mouseMode'] != pg.ViewBox.RectMode or ev.button() == QtCore.Qt.RightButton:
        #         ev.ignore()

        #     if ev.isFinish():  ## This is the final move in the drag; draw the actual box
        #         print("dragging done")
        #         self.rbScaleBox.hide()
        #         self.sigMouseDragged.emit(ev.buttonDownScenePos(ev.button()),ev.scenePos(),ev.screenPos())
        #     else:
        #         ## update shape of scale box
        #         self.updateScaleBox(ev.buttonDownPos(), ev.pos())
        # else:
        #     pass

    def mousePressEvent(self, ev):
        if self.enableDrag and ev.button() == self.parent.MouseDrawingButton:
            if self.thisIsAmpl:
                self.parent.mouseClicked_ampl(ev)
            else:
                self.parent.mouseClicked_spec(ev)
            ev.accept()
        else:
            ev.ignore()

    def mouseReleaseEvent(self, ev):
        if self.enableDrag and ev.button() == self.parent.MouseDrawingButton:
            if self.thisIsAmpl:
                self.parent.mouseClicked_ampl(ev)
            else:
                self.parent.mouseClicked_spec(ev)
            ev.accept()
        else:
            ev.ignore()

    def keyPressEvent(self,ev):
        # This catches the keypresses and sends out a signal
        #self.emit(SIGNAL("keyPressed"),ev)
        super(DragViewBox, self).keyPressEvent(ev)
        self.keyPressed.emit(ev.key())

class ChildInfoViewBox(pg.ViewBox):
    # Normal ViewBox, but with ability to pass a message back from a child
    sigChildMessage = QtCore.Signal(object)

    def __init__(self, *args, **kwds):
        pg.ViewBox.__init__(self, *args, **kwds)

    def resend(self,x):
        self.sigChildMessage.emit(x)

class ClickableRectItem(QtGui.QGraphicsRectItem):
    # QGraphicsItem doesn't include signals, hence this mess
    def __init__(self, *args, **kwds):
        QtGui.QGraphicsRectItem.__init__(self, *args, **kwds)

    def mousePressEvent(self, ev):
        super(ClickableRectItem, self).mousePressEvent(ev)
        self.parentWidget().resend(self.mapRectToParent(self.boundingRect()).x())

class ControllableAudio(QAudioOutput):
    # This links all the PyQt5 audio playback things -
    # QAudioOutput, QFile, and input from main interfaces

    def __init__(self, format):
        super(ControllableAudio, self).__init__(format)
        # on this notify, move slider (connected in main file)
        self.setNotifyInterval(30)
        self.stateChanged.connect(self.endListener)
        self.tempin = QBuffer()
        self.startpos = 0
        self.timeoffset = 0
        self.keepSlider = False
        self.format = format
        # set small buffer (10 ms) and use processed time
        self.setBufferSize(int(self.format.sampleSize() * self.format.sampleRate()/100 * self.format.channelCount()))

    def isPlaying(self):
        return(self.state() == QAudio.ActiveState)

    def endListener(self):
        # this should only be called if there's some misalignment between GUI and Audio
        if self.state() == QAudio.IdleState:
            # give some time for GUI to catch up and stop
            sleepCycles = 0
            while(self.state() != QAudio.StoppedState and sleepCycles < 30):
                sleep(0.03)
                sleepCycles += 1
                # This loop stops when timeoffset+processedtime > designated stop position.
                # By adding this offset, we ensure the loop stops even if
                # processed audio timer breaks somehow.
                self.timeoffset += 30
                self.notify.emit()
            self.pressedStop()

    def pressedPlay(self, resetPause=False, start=0, stop=0, audiodata=None):
        if not resetPause and self.state() == QAudio.SuspendedState:
            print("Resuming at: %d" % self.pauseoffset)
            self.sttime = time.time() - self.pauseoffset/1000
            self.resume()
        else:
            if not self.keepSlider or resetPause:
                self.pressedStop()

            print("Starting at: %d" % self.tempin.pos())
            sleep(0.2)
            # in case bar was moved under pause, we need this:
            pos = self.tempin.pos() # bytes
            pos = self.format.durationForBytes(pos) / 1000 # convert to ms
            pos = pos + start
            print("Pos: %d start: %d stop %d" %(pos, start, stop))
            self.filterSeg(pos, stop, audiodata)

    def pressedPause(self):
        self.keepSlider=True # a flag to avoid jumping the slider back to 0
        pos = self.tempin.pos() # bytes
        pos = self.format.durationForBytes(pos) / 1000 # convert to ms
        # store offset, relative to the start of played segment
        self.pauseoffset = pos + self.timeoffset
        self.suspend()

    def pressedStop(self):
        # stop and reset to window/segment start
        self.keepSlider=False
        self.stop()
        if self.tempin.isOpen():
            self.tempin.close()

    def filterBand(self, start, stop, low, high, audiodata, sp):
        # takes start-end in ms, relative to file start
        self.timeoffset = max(0, start)
        start = max(0, start * self.format.sampleRate() // 1000)
        stop = min(stop * self.format.sampleRate() // 1000, len(audiodata))
        segment = audiodata[int(start):int(stop)]
        segment = sp.bandpassFilter(segment,sampleRate=None, start=low, end=high)
        # segment = self.sp.ButterworthBandpass(segment, self.sampleRate, bottom, top,order=5)
        self.loadArray(segment)

    def filterSeg(self, start, stop, audiodata):
        # takes start-end in ms
        self.timeoffset = max(0, start)
        start = max(0, int(start * self.format.sampleRate() // 1000))
        stop = min(int(stop * self.format.sampleRate() // 1000), len(audiodata))
        segment = audiodata[start:stop]
        self.loadArray(segment)

    def loadArray(self, audiodata):
        # loads an array from memory into an audio buffer
        if self.format.sampleSize() == 16:
            audiodata = audiodata.astype('int16')  # 16 corresponds to sampwidth=2
        elif self.format.sampleSize() == 32:
            audiodata = audiodata.astype('int32')
        elif self.format.sampleSize() == 24:
            audiodata = audiodata.astype('int32')
            print("Warning: 24-bit sample playback currently not supported")
        elif self.format.sampleSize() == 8:
            audiodata = audiodata.astype('uint8')
        else:
            print("ERROR: sampleSize %d not supported" % self.format.sampleSize())
            return
        # double mono sound to get two channels - simplifies reading
        if self.format.channelCount()==2:
            audiodata = np.column_stack((audiodata, audiodata))

        # write filtered output to a BytesIO buffer
        self.tempout = io.BytesIO()
        # NOTE: scale=None rescales using data minimum/max. This can cause clipping. Use scale="none" if this causes weird playback sound issues.
        # in particular for 8bit samples, we need more scaling:
        if self.format.sampleSize() == 8:
            scale = (audiodata.min()/2, audiodata.max()*2)
        else:
            scale = None
        wavio.write(self.tempout, audiodata, self.format.sampleRate(), scale=scale, sampwidth=self.format.sampleSize() // 8)

        # copy BytesIO@write to QBuffer@read for playing
        self.temparr = QByteArray(self.tempout.getvalue()[44:])
        # self.tempout.close()
        if self.tempin.isOpen():
            self.tempin.close()
        self.tempin.setBuffer(self.temparr)
        self.tempin.open(QIODevice.ReadOnly)

        # actual timer is launched here, with time offset set asynchronously
        sleep(0.2)
        self.sttime = time.time() - self.timeoffset/1000
        self.start(self.tempin)

    def seekToMs(self, ms, start):
        print("Seeking to %d ms" % ms)
        # start is an offset for the current view start, as it is position 0 in extracted file
        self.reset()
        self.tempin.seek(self.format.bytesForDuration((ms-start)*1000))
        self.timeoffset = ms

    def applyVolSlider(self, value):
        # passes UI volume nonlinearly
        # value = QAudio.convertVolume(value / 100, QAudio.LogarithmicVolumeScale, QAudio.LinearVolumeScale)
        value = (math.exp(value/50)-1)/(math.exp(2)-1)
        self.setVolume(value)

class FlowLayout(QtGui.QLayout):
    # This is the flow layout which lays out a set of spectrogram pictures on buttons (for HumanClassify2) as
    # nicely as possible
    # From https://gist.github.com/Cysu/7461066
    def __init__(self, parent=None, margin=0, spacing=-1):
        super(FlowLayout, self).__init__(parent)

        if parent is not None:
            self.setMargin(margin)

        self.setSpacing(spacing)

        self.itemList = []

        self.margin = margin

    def __del__(self):
        item = self.takeAt(0)
        while item:
            item = self.takeAt(0)

    def addItem(self, item):
        self.itemList.append(item)

    def count(self):
        return len(self.itemList)

    def itemAt(self, index):
        if index >= 0 and index < len(self.itemList):
            return self.itemList[index]

        return None

    def takeAt(self, index):
        if index >= 0 and index < len(self.itemList):
            return self.itemList.pop(index)

        return None

    def expandingDirections(self):
        return QtCore.Qt.Orientations(QtCore.Qt.Orientation(0))

    def hasHeightForWidth(self):
        return True

    def heightForWidth(self, width):
        height = self._doLayout(QtCore.QRect(0, 0, width, 0), True)
        return height

    def setGeometry(self, rect):
        super(FlowLayout, self).setGeometry(rect)
        self._doLayout(rect, False)

    def sizeHint(self):
        return self.minimumSize()

    # def minimumSize(self):
    #     size = QtCore.QSize()
    #
    #     for item in self.itemList:
    #         size = size.expandedTo(item.minimumSize())
    #
    #     size += QtCore.QSize(2 * self.margin(), 2 * self.margin())
    #     return size

    def _doLayout(self, rect, testOnly):
        x = rect.x()
        y = rect.y()
        lineHeight = 0

        for item in self.itemList:
            wid = item.widget()
            spaceX = self.spacing() + wid.style().layoutSpacing(
                QtGui.QSizePolicy.PushButton,
                QtGui.QSizePolicy.PushButton,
                QtCore.Qt.Horizontal)

            spaceY = self.spacing() + wid.style().layoutSpacing(
                QtGui.QSizePolicy.PushButton,
                QtGui.QSizePolicy.PushButton,
                QtCore.Qt.Vertical)

            nextX = x + item.sizeHint().width() + spaceX
            if nextX - spaceX > rect.right() and lineHeight > 0:
                x = rect.x()
                y = y + lineHeight + spaceY
                nextX = x + item.sizeHint().width() + spaceX
                lineHeight = 0

            if not testOnly:
                item.setGeometry(
                    QtCore.QRect(QtCore.QPoint(x, y), item.sizeHint()))

            x = nextX
            lineHeight = max(lineHeight, item.sizeHint().height())

        return y + lineHeight - rect.y()

class Log(object):
    """ Used for logging info during batch processing.
        Stores most recent analysis for each species, to stay in sync w/ data files.
        Arguments:
        1. path to log file
        2. species
        3. list of other settings of the current analysis

        LOG FORMAT, for each analysis:
        #freetext line
        species
        settings line
        files, multiple lines
    """

    def __init__(self, path, species, settings):
        # in order to append, the previous log must:
        # 1. exist
        # 2. be writeable
        # 3. match current analysis
        # On init, we parse the existing log to see if appending is possible.
        # Actual append/create happens later.
        self.possibleAppend = False
        self.file = path
        self.species = species
        self.settings = ','.join(map(str, settings))
        self.oldAnalyses = []
        self.filesDone = []
        self.currentHeader = ""
        allans = []

        # now, check if the specified log can be resumed:
        if os.path.isfile(path):
            try:
                f = open(path, 'r+')
                print("Found log file at %s" % path)

                lines = [line.rstrip('\n') for line in f]
                f.close()
                lstart = 0
                lend = 1
                # parse to separate each analysis into
                # [freetext, species, settings, [files]]
                # (basically I'm parsing txt into json because I'm dumb)
                while lend<len(lines):
                    #print(lines[lend])
                    if lines[lend][0] == "#":
                        allans.append([lines[lstart], lines[lstart+1], lines[lstart+2],
                                        lines[lstart+3 : lend]])
                        lstart = lend
                    lend += 1
                allans.append([lines[lstart], lines[lstart+1], lines[lstart+2],
                                lines[lstart+3 : lend]])

                # parse the log thusly:
                # if current species analysis found, store parameters
                # and compare to check if it can be resumed.
                # store all other analyses for re-printing.
                for a in allans:
                    #print(a)
                    if a[1]==self.species:
                        print("Resumable analysis found")
                        # do not reprint this in log
                        if a[2]==self.settings:
                            self.currentHeader = a[0]
                            # (a1 and a2 match species & settings anyway)
                            self.filesDone = a[3]
                            self.possibleAppend = True
                    else:
                        # store this for re-printing to log
                        self.oldAnalyses.append(a)

            except IOError:
                # bad error: lacking permissions?
                print("ERROR: could not open log at %s" % path)

    def appendFile(self, filename):
        print('Appending %s to log' % filename)
        # attach file path to end of log
        self.file.write(filename)
        self.file.write("\n")
        self.file.flush()

    def appendHeader(self, header, species, settings):
        if header is None:
            header = "#Analysis started on " + time.strftime("%Y %m %d, %H:%M:%S") + ":"
        self.file.write(header)
        self.file.write("\n")
        self.file.write(species)
        self.file.write("\n")
        if type(settings) is list:
            settings = ','.join(settings)
        self.file.write(settings)
        self.file.write("\n")
        self.file.flush()

    def reprintOld(self):
        # push everything from oldAnalyses to log
        # To be called once starting a new log is confirmed
        for a in self.oldAnalyses:
            self.appendHeader(a[0], a[1], a[2])
            for f in a[3]:
                self.appendFile(f)


class MessagePopup(QMessageBox):
    """ Convenience wrapper around QMessageBox.
        TYPES, based on main icon:
        w - warning
        d - done (successful completion)
        t - thinking (questions)
        o - other
        a - about
    """
    def __init__(self, type, title, text):
        super(QMessageBox, self).__init__()

        self.setText(text)
        self.setWindowTitle(title)
        if (type=="w"):
            self.setIconPixmap(QPixmap("img/Owl_warning.png"))
        elif (type=="d"):
            self.setIcon(QMessageBox.Information)
            self.setIconPixmap(QPixmap("img/Owl_done.png"))
        elif (type=="t"):
            self.setIcon(QMessageBox.Information)
            self.setIconPixmap(QPixmap("img/Owl_thinking.png"))
        elif (type=="a"):
            # Easy way to set ABOUT text here:
            self.setIconPixmap(QPixmap("img/AviaNZ.png"))
            self.setText("The AviaNZ Program, v1.5 (August 2019)")
            self.setInformativeText("By Stephen Marsland, Victoria University of Wellington. With code by Nirosha Priyadarshani and Julius Juodakis, and input from Isabel Castro, Moira Pryde, Stuart Cockburn, Rebecca Stirnemann, Sumudu Purage, Virginia Listanti, and Rebecca Huistra. \n stephen.marsland@vuw.ac.nz")
        elif (type=="o"):
            self.setIconPixmap(QPixmap("img/AviaNZ.png"))

        self.setWindowIcon(QIcon("img/Avianz.ico"))

        # by default, adding OK button. Can easily be overwritten after creating
        self.setStandardButtons(QMessageBox.Ok)


class ConfigLoader(object):
    """ This deals with reading main config files.
        Not much functionality, but lots of exception handling,
        so moved it out separately.

        Most of these functions return the contents of a corresponding JSON file.
    """

    def config(self, file):
        # At this point, the main config file should already be ensured to exist.
        # It will always be in user configdir, otherwise it would be impossible to find.
        print("Loading software settings from file %s" % file)
        try:
            config = json.load(open(file))
            return config
        except ValueError as e:
            # if JSON looks corrupt, quit:
            print(e)
            msg = MessagePopup("w", "Bad config file", "ERROR: file " + file + " corrupt, delete it to restore default")
            msg.exec_()
            sys.exit()

    def filters(self, dir):
        # Returns a list of filter files.
        print("Loading call filters from folder %s" % dir)
        try:
            filters = [f[:-4] for f in os.listdir(dir) if os.path.isfile(os.path.join(dir, f))]
            return filters
        except Exception:
            print("Folder %s not found, no filters loaded" % dir)
            return None

    def shortbl(self, file, configdir):
        # A fallback shortlist will be confirmed to exist in configdir.
        # This list is necessary, long list can be None
        print("Loading short species list from file %s" % file)
        try:
            if os.path.isabs(file):
                # user-picked files will have absolute paths
                shortblfile = file
            else:
                # initial file will have relative path,
                # to allow looking it up in various OSes.
                shortblfile = os.path.join(configdir, file)
            if not os.path.isfile(shortblfile):
                print("Warning: file %s not found, falling back to default" % shortblfile)
                shortblfile = os.path.join(configdir, "ListCommonBirds.txt")

            try:
                readlist = json.load(open(shortblfile))
                return readlist
            except ValueError as e:
                # if JSON looks corrupt, quit and suggest deleting:
                print(e)
                msg = MessagePopup("w", "Bad species list", "ERROR: file " + shortblfile + " corrupt, delete it to restore default")
                msg.exec_()
                return None

        except Exception as e:
            # if file is not found at all, quit, user must recreate the file or change path
            print(e)
            msg = MessagePopup("w", "Bad species list", "ERROR: Failed to load short species list from " + file)
            msg.exec_()
            return None


    def longbl(self, file, configdir):
        if file == "None":
            # long bird list can be set to "None" intentionally
            return None

        print("Loading long species list from file %s" % file)
        try:
            if os.path.isabs(file):
                # user-picked files will have absolute paths
                longblfile = file
            else:
                # initial file will have relative path,
                # to allow looking it up in various OSes.
                longblfile = os.path.join(configdir, file)
            if not os.path.isfile(longblfile):
                print("Warning: file %s not found, falling back to default" % longblfile)
                longblfile = os.path.join(configdir, "ListDOCBirds.txt")

            try:
                readlist = json.load(open(longblfile))
                return readlist
            except ValueError as e:
                print(e)
                msg = MessagePopup("w", "Bad species list", "Warning: file " + longblfile + " corrupt, delete it to restore default")
                msg.exec_()
                return None

        except Exception as e:
            print(e)
            msg = MessagePopup("w", "Bad species list", "Warning: Failed to load long species list from " + file)
            msg.exec_()
            return None

    # Dumps the provided JSON array to the corresponding bird file.
    def blwrite(self, content, file, configdir):
        print("Updating species list in file %s" % file)
        try:
            if os.path.isabs(file):
                # user-picked files will have absolute paths
                file = file
            else:
                # initial file will have relative path,
                # to allow looking it up in various OSes.
                file = os.path.join(configdir, file)

            # no fallback in case file not found - don't want to write to random places.
            json.dump(content, open(file, 'w'), indent=1)

        except Exception as e:
            print(e)
            msg = MessagePopup("w", "Unwriteable species list", "Warning: Failed to write species list to " + file)
            msg.exec_()

    # Dumps the provided JSON array to the corresponding config file.
    def configwrite(self, content, file):
        print("Saving config to file %s" % file)
        try:
            # will always be an absolute path to the user configdir.
            json.dump(content, open(file, 'w'), indent=1)

        except Exception as e:
            print("ERROR while saving config file:")
            print(e)
